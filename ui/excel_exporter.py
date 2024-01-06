import pandas as pd
import openpyxl
import xlsxwriter

import os


class ExcelExporter:
    def __init__(self, excel_dir_path: str):
        self.__excel_dir_path = excel_dir_path
        self.__excel_file_path = self.__excel_dir_path + "/summary_export.xlsx"
        self.__classification_excel_path = excel_dir_path + "/classification_export.xlsx"
        self.__descriptive_statistics_excel_path = excel_dir_path + "/descriptive_statistics.xlsx"
        self.__create_excel_dir()
        self.__similarity_classification_criteria = ["Peak_hour", "Peak_period_1_hour_a", "Peak_period_2_hour_a",
                                                     "Peak_period_1_hour_b"]
        self.__similarity_classification_criteria_index = 0

    def set_similarity_criteria_index(self, new_index):
        self.__similarity_classification_criteria_index = new_index
        print("New index is set for similarity index", new_index)

    def write(self, export_columns_dict):

        # with pd.ExcelWriter(self.__excel_file_path) as writer:
        #     # writing to the 'Employee' sheet
        #     data_frame = pd.DataFrame(export_columns_dict)
        #     data_frame.to_excel(writer, sheet_name='Summary', index=False)
        #     worksheet = writer.sheets["Summary"]
        #     worksheet.set_column('segment_id:', 70)
        if os.path.exists(self.__excel_file_path):
            os.remove(self.__excel_file_path)

        workbook = xlsxwriter.Workbook(self.__excel_file_path)
        worksheet = workbook.add_worksheet('Summary')
        cell_format = workbook.add_format({'num_format': '#,##0'})
        cell_format.set_align('center')
        cell_format.set_align('vcenter')

        col_num = 0
        for key, value in export_columns_dict.items():
            worksheet.set_row(0, 20)
            worksheet.set_column(col_num, col_num, 25)
            worksheet.write(0, col_num, key, cell_format)
            worksheet.write_column(1, col_num, value, cell_format)
            col_num += 1

        workbook.close()

        self.__similarity_classification(export_columns_dict)
        self.__descriptive_statistics(export_columns_dict)

        print('DataFrames are written to Excel File successfully.')

    def __create_excel_dir(self):
        if not os.path.exists(self.__excel_dir_path):
            os.mkdir(self.__excel_dir_path)
            print("Excel file directory in : ", self.__excel_dir_path)
        else:
            print("Excel file directory exists already : ", self.__excel_dir_path)

    def __descriptive_statistics(self, export_columns_dict):

        time_sets_column = export_columns_dict["Peak_hour"]
        total = len(time_sets_column)
        time_sets_dict = {}
        for time_set in time_sets_column:

            if time_set not in list(time_sets_dict.keys()):
                time_sets_dict[time_set] = 0
                time_sets_dict[time_set] = time_sets_dict[time_set]+1
            else :
                time_sets_dict[time_set] = time_sets_dict[time_set] + 1

        workbook = xlsxwriter.Workbook(self.__descriptive_statistics_excel_path)

        cell_format = workbook.add_format({'num_format': '#,##0'})
        float_cell_format = workbook.add_format()
        cell_format.set_align('center')
        cell_format.set_align('vcenter')
        float_cell_format.set_align('center')
        float_cell_format.set_align('vcenter')

        worksheet = workbook.add_worksheet()
        worksheet.set_row(0, 20)

        worksheet.write(0, 0, "TimeSet", cell_format)
        worksheet.write(0, 1, "Count", cell_format)
        worksheet.write(0, 2, "Percent", cell_format)
        worksheet.write(0, 3, "Total", cell_format)
        worksheet.write(1, 3, total, cell_format)

        row_num = 1
        for peak_hour in time_sets_dict.keys():

            count = time_sets_dict[peak_hour]
            percent = count/total * 100


            # width of row
            worksheet.set_row(row_num, 20)

            # height of col
            worksheet.set_column(0, 3, 25)

            worksheet.write(row_num, 0, peak_hour, cell_format)
            worksheet.write(row_num, 1, count, cell_format)
            worksheet.write(row_num, 2, percent, float_cell_format)

            row_num += 1

        workbook.close()



    def __similarity_classification(self, export_columns_dict):

        similarity_classification_column = self.__similarity_classification_criteria[
            self.__similarity_classification_criteria_index]

        similarity_criteria = export_columns_dict[similarity_classification_column]
        address_column = export_columns_dict['Address']
        peak_hour_class_dict = {}

        for index, peak_hour in enumerate(similarity_criteria):
            peak_hour_class_dict[str(peak_hour)] = []

        for index, peak_hour in enumerate(similarity_criteria):
            peak_hour_class_dict[str(peak_hour)].append(address_column[index])

        workbook = xlsxwriter.Workbook(self.__classification_excel_path)

        cell_format = workbook.add_format({'num_format': '#,##0'})
        cell_format.set_align('center')
        cell_format.set_align('vcenter')

        for peak_hour in peak_hour_class_dict.keys():
            values = peak_hour_class_dict[peak_hour]
            worksheet = workbook.add_worksheet()

            col_num = 0

            worksheet.set_row(0, 20)
            worksheet.set_column(col_num, col_num, 25)
            worksheet.write(0, col_num, peak_hour, cell_format)
            worksheet.write_column(1, col_num, values, cell_format)
            col_num += 1

        workbook.close()
